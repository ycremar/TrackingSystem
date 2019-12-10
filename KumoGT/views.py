from django.shortcuts import render, redirect, get_object_or_404
from django.conf import settings
from django.core.files.storage import FileSystemStorage
from django.http import FileResponse, HttpResponse, Http404, HttpResponseRedirect
from django.contrib import messages
from django.utils.safestring import mark_safe

from django.contrib.auth.decorators import login_required, user_passes_test
from django.views.static import serve

from .models import Deg_Plan_Doc, Student, Degree, Pre_Exam_Doc, Pre_Exam_Info,\
    T_D_Prop_Doc, Fin_Exam_Info, Fin_Exam_Doc, T_D_Doc, T_D_Info, Session_Notes,\
    Other_Doc
from django.core.paginator import Paginator

from .forms import create_doc_form, stu_search_form, stu_bio_form, deg_form,\
    pre_exam_info_form, final_exam_info_form, thesis_dissertation_info_form, session_notes_form
from .crypt import Cryptographer
from .functions import deg_doc, get_info_form, get_stu_objs, post_degrees, post_session_notes,\
    delete, get_stu_search_dict

from openpyxl import Workbook

import os

def conditional_decorator(dec, condition):
    def decorator(func):
        if not condition:
            return func
        return dec(func)
    return decorator

@conditional_decorator(login_required(login_url='/login/'), not settings.DEBUG)
def home(request):
    docs = Deg_Plan_Doc.objects.all()
    return render(request, 'home.html', { 'docs': docs })

@conditional_decorator(login_required(login_url='/login/'), not settings.DEBUG)
def upload(request):
    if request.method == 'POST' and request.FILES['myfile']:
        myfile = request.FILES['myfile']
        fs = FileSystemStorage()
        filename = fs.save(myfile.name, myfile)
        uploaded_file_url = fs.url(filename)
        return render(request, 'upload.html', {
            'uploaded_file_url': uploaded_file_url
        })
    return render(request, 'upload.html')
    
@conditional_decorator(login_required(login_url='/login/'), not settings.DEBUG)
def form_upload(request):
    if request.method == 'POST':
        form = create_doc_form(Deg_Plan_Doc)(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            return redirect('home')
    else:
        form = create_doc_form(Deg_Plan_Doc)
    return render(request, 'form_upload.html', {
        'form': form
    })

@conditional_decorator(login_required(login_url='/login/'), not settings.DEBUG)
def degree_plan(request, deg_id, option = '', id = 0):
    if request.method == 'POST':
        return deg_doc(request, "Degree Plan", Deg_Plan_Doc, "/degree_plan/", deg_id, option, id)[1]
    else:
        method, data = deg_doc(request, "Degree Plan", Deg_Plan_Doc, "/degree_plan/", deg_id, option, id)
        if method != 'show': return data
        else:
            deg, forms = data
            return render(request, 'degree_plan.html', {
                'deg': deg,
                'forms': forms,
                'option': option,
            })

@conditional_decorator(login_required(login_url='/login/'), not settings.DEBUG)
def preliminary_exam(request, deg_id, option = '', id = 0):
    info_form = get_info_form(request, deg_id, Pre_Exam_Info, pre_exam_info_form)
    if request.method == 'POST':
        return deg_doc(request, "Preliminary Exam", Pre_Exam_Doc, "/preliminary_exam/",\
            deg_id, option, id, Pre_Exam_Info, info_form)[1]
    else:
        method, data = deg_doc(request, "Preliminary Exam", Pre_Exam_Doc, "/preliminary_exam/",\
            deg_id, option, id, Pre_Exam_Info)
        if method != 'show': return data
        else:
            deg, forms = data
            return render(request, 'preliminary_exam.html', {
                'deg': deg,
                'forms': forms,
                'option': option,
                'info_form': info_form,
            })

@conditional_decorator(login_required(login_url='/login/'), not settings.DEBUG)
def thesis_dissertation_proposal(request, deg_id, option = '', id = 0):
    if request.method == 'POST':
        return deg_doc(request, "Thesis/Dissertation Proposal", T_D_Prop_Doc, "/thesis_dissertation_proposal/",\
            deg_id, option, id)[1]
    else:
        method, data = deg_doc(request, "Thesis/Dissertation Proposal", T_D_Prop_Doc, "/thesis_dissertation_proposal/",\
            deg_id, option, id)
        if method != 'show': return data
        else:
            deg, forms = data
            return render(request, 'thesis_dissertation_proposal.html', {
                'deg': deg,
                'forms': forms,
                'option': option,
            })

@conditional_decorator(login_required(login_url='/login/'), not settings.DEBUG)
def final_exam(request, deg_id, option = '', id = 0):
    info_form = get_info_form(request, deg_id, Fin_Exam_Info, final_exam_info_form)
    if request.method == 'POST':
        return deg_doc(request, "Final Exam", Fin_Exam_Doc, "/final_exam/",\
            deg_id, option, id, Fin_Exam_Info, info_form)[1]
    else:
        method, data = deg_doc(request, "Final Exam", Fin_Exam_Doc, "/final_exam/",\
            deg_id, option, id, Fin_Exam_Info)
        if method != 'show': return data
        else:
            deg, forms = data
            return render(request, 'final_exam.html', {
                'deg': deg,
                'forms': forms,
                'option': option,
                'info_form': info_form,
            })
            
@conditional_decorator(login_required(login_url='/login/'), not settings.DEBUG)
def thesis_dissertation(request, deg_id, option = '', id = 0):
    info_form = get_info_form(request, deg_id, T_D_Info, thesis_dissertation_info_form)
    if request.method == 'POST':
        return deg_doc(request, "Thesis/Dissertation", T_D_Doc, "/thesis_dissertation/",\
            deg_id, option, id, T_D_Info, info_form)[1]
    else:
        method, data = deg_doc(request, "Thesis/Dissertation", T_D_Doc, "/thesis_dissertation/",\
            deg_id, option, id, T_D_Info)
        if method != 'show': return data
        else:
            deg, forms = data
            return render(request, 'thesis_dissertation.html', {
                'deg': deg,
                'forms': forms,
                'option': option,
                'info_form': info_form,
            })

@conditional_decorator(login_required(login_url='/login/'), not settings.DEBUG)
def other_doc(request, deg_id, option = '', id = 0):
    if request.method == 'POST':
        return deg_doc(request, "Other Document", Other_Doc, "/other_doc/", deg_id, option, id)[1]
    else:
        method, data = deg_doc(request, "Other Document", Other_Doc, "/other_doc/", deg_id, option, id)
        if method != 'show': return data
        else:
            deg, forms = data
            return render(request, 'other_doc.html', {
                'deg': deg,
                'forms': forms,
                'option': option,
            })

@conditional_decorator(login_required(login_url='/login/'), not settings.DEBUG)
def session_notes(request, stu_id, option = '', id = 0):
    if option == 'del':
        return delete(request, Session_Notes, id, "Session Note", "advised at",\
            "date", "/student/" + stu_id + "/session_notes/")
    if request.method == 'POST':
        return post_session_notes(request, stu_id, option, id)
    else:
        notes, student, new_form = get_stu_objs(Session_Notes, session_notes_form, stu_id, option, False)
        return render(request, 'session_notes.html', {
            'stu': student,
            'notes': notes,
            'new_form': new_form,
            'option': option,
        })

@conditional_decorator(login_required(login_url='/login/'), not settings.DEBUG)
def serve_protected_document(request, file_path):
    try:
        if file_path[0:6] != 'media/' or '../' in file_path: raise PermissionError
        file_path = os.path.join(settings.BASE_DIR, file_path)
        with open(file_path, 'rb') as fh:
            content = Cryptographer.decrypted(fh.read())
            response = HttpResponse(content, content_type="application/pdf")
            response['Content-Disposition'] = 'inline; filename=' + os.path.basename(file_path)
            return response
    except:
        raise Http404

@conditional_decorator(login_required(login_url='/login/'), not settings.DEBUG)
def students(request, **kwargs):# uin, first_name, last_name, gender, status, cur_degree):
    if request.method == 'POST':
        form = stu_search_form(request.POST)
        form.is_valid()
        redirect_url = '/students/'
        search_form_params = {}
        for name, val in form.cleaned_data.items():
            if val and val != '':
                search_form_params[name] = val
                redirect_url += "{0}={1}/".format(name, val)
        return redirect(redirect_url, **search_form_params)
    else:
        students = Student.objects.all()
        seach_dict, search_form_params = get_stu_search_dict(kwargs, True)
        if kwargs: students = students.filter(**seach_dict)
        students = students.order_by('uin')
        form = stu_search_form(search_form_params)
        paginator = Paginator(students, 20) # Show 20 students per page. Use 1 for test.
        page = request.GET.get('page')
        students_page = paginator.get_page(page)
        page = 1 if not page else int(page)
        neigh_pages = [n for n in range(max(page - 2, 1), min(page + 3, paginator.num_pages + 1))]
        if len(neigh_pages) == 0 or neigh_pages[0] > 1:
            neigh_pages.insert(0, -1)
            neigh_pages.insert(0, 1)
        if neigh_pages[-1] < paginator.num_pages:
            neigh_pages.append(-1)
            neigh_pages.append(paginator.num_pages)
    return render(request, 'students.html', {
        'form': form,
        'students': students_page,
        'neigh_pages': neigh_pages,
        })
        
@conditional_decorator(login_required(login_url='/login/'), not settings.DEBUG)
def create_stu(request, back_url = None):
    if request.method == 'POST':
        form = stu_bio_form(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Student is added.')
        else: 
            messages.error(request, mark_safe("{0}".format(form.errors)))
        return redirect('create_stu')
    else:
        form = stu_bio_form()
        title = 'Add a Student'
        return render(request, 'stu_bio_info.html', {
            'form': form,
            'title': title,
            })
            
@conditional_decorator(login_required(login_url='/login/'), not settings.DEBUG)
def edit_stu(request, id, back_url = None):
    if request.method == 'POST':
        form = stu_bio_form(request.POST, instance = Student.objects.get(id = id))
        if form.has_changed():
            if form.is_valid():
                form.save()
                messages.success(request, 'Student is updated.')
            else: 
                messages.error(request, mark_safe("{0}".format(form.errors)))
        else:
            messages.info(request, 'Noting is changed.')
        return redirect('edit_stu', id = id)
    else:
        form = stu_bio_form(instance = Student.objects.get(id = id))
        title = 'Edit Student Bio Info'
        return render(request, 'stu_bio_info.html', {
            'form': form,
            'title': title,
            })
            
@conditional_decorator(login_required(login_url='/login/'), not settings.DEBUG)
def delete_stu(request, id):
    return delete(request, Student, id, "Student", 'UIN', 'uin', '/students/')

@conditional_decorator(login_required(login_url='/login/'), not settings.DEBUG)
def degrees(request, stu_id, option = '', id = 0):
    if option == 'del':
        return delete(request, Degree, id, "Degree", "",\
            "deg_type", "/student/" + stu_id + "/degrees/")
    if request.method == 'POST':
        return post_degrees(request, stu_id, option, id)
    else:
        forms, student = get_stu_objs(Degree, deg_form, stu_id, option)
        cur_deg_id = student.cur_degree.id if student and student.cur_degree else -1
        return render(request, 'degrees.html', {
            'stu': student,
            'cur_deg_id': cur_deg_id,
            'forms': forms,
            'option': option,
        })

@user_passes_test(lambda u: u.is_superuser)
def download_stu_info(request, **kwargs):
    models = [Student, Degree, Pre_Exam_Info, T_D_Info]
    if request.method == 'POST':
        fields = {}
        wb = Workbook()
        ws = wb.active
        ws.title = "Information of Students"
        i, j = 1, 1
        for model in models:
            name = model._meta.model_name
            sub_fields = request.POST.getlist(name)
            sub_fields_display = []
            for field in model._meta.fields:
                prefix = ""
                if name != 'student' and name != 'degree': prefix = model._meta.verbose_name + " "
                if field.attname in sub_fields:
                    sub_fields_display.append(prefix + field.verbose_name)
            for field in sub_fields_display:
                ws.cell(row = i, column = j, value = field)
                j += 1
            fields[model._meta.model_name] = sub_fields
        i += 1
        students = Student.objects.all()
        seach_dict = get_stu_search_dict(kwargs)
        if kwargs: students = students.filter(**seach_dict)
        students = students.order_by('uin')
        for stu in students:
            j = 1
            for model in models:
                name = model._meta.model_name
                if name == 'student':
                    obj = stu
                elif name == 'degree':
                    obj = stu.cur_degree
                else:
                    obj = getattr(stu.cur_degree, name) if hasattr(stu.cur_degree, name) else None
                if obj:
                    for field in fields[name]:
                        if hasattr(obj, "get_{0}_display".format(field)):
                            text = getattr(obj, "get_{0}_display".format(field))()
                            ws.cell(row = i, column = j, value = text)
                        else:
                            ws.cell(row = i, column = j, value = obj.__dict__[field])
                        j += 1
                else: j += len(fields[name])
            i += 1
        file_name = "data.xlsx"
        file_path = os.path.join(settings.MEDIA_ROOT, file_name)
        try:
            if not os.path.isdir(settings.MEDIA_ROOT):
                os.mkdir(settings.MEDIA_ROOT)
            wb.save(file_path)
        except:
            messages.error(request, "Some error has occured, please retry it later.")
            return redirect('download_stu_info')
        finally:
            return redirect('get_tmp_file',\
                file_path = file_name,\
                content_type = "application/vnd.ms-excel")
    else:
        fields = []
        for model in models:
            sub_fields = []
            for field in model._meta.fields:
                if not 'id' in field.attname:
                    sub_fields.append((field.attname, field.verbose_name))
            fields.append([model._meta.model_name, model._meta.verbose_name, sub_fields])
        return render(request, 'download_stu_info.html', {
            'fields': fields,
        })

class Tmp_File(object): 
    def __init__(self, filename): 
        self.filename = filename 
        self.fd = None
        
    def open(self, mode): 
        if self.fd == None: 
            self.fd = open(self.filename, mode) 
        return self.fd 

    def __del__(self): 
        if self.fd != None: 
            self.fd.close() 
            os.remove(self.filename) 

@conditional_decorator(login_required(login_url='/login/'), not settings.DEBUG)
def get_tmp_file(request, file_path, content_type):
    try:
        if file_path[0:10] == 'documents/' or '../' in file_path: raise PermissionError
        file_path = os.path.join(settings.MEDIA_ROOT, file_path)
        fh = Tmp_File(file_path)
        content = fh.open('rb')
        response = HttpResponse(content, content_type = content_type)
        response['Content-Disposition'] = 'attachment; filename=' + os.path.basename(file_path)
        return response
    except:
        raise Http404